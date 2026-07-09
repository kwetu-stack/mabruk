import os

from openpyxl import load_workbook

from models import db
from models.supplier import Supplier
from models.product import Product


class ImportService:

    @staticmethod
    def resolve_file_path(file_path):
        if os.path.isabs(file_path):
            return file_path

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(project_root, file_path)

    @staticmethod
    def load_excel(file_path):
        """
        Load an Excel workbook.
        """
        return load_workbook(ImportService.resolve_file_path(file_path))

    @staticmethod
    def import_suppliers(file_path):

        workbook = load_workbook(ImportService.resolve_file_path(file_path))

        created = 0
        skipped = 0

        for sheet in workbook.sheetnames:

            supplier_name = sheet.strip()

            supplier = Supplier.query.filter_by(name=supplier_name).first()

            if supplier:
                skipped += 1
                continue

            new_supplier = Supplier(name=supplier_name)

            db.session.add(new_supplier)

            created += 1

        db.session.commit()

        return {"created": created, "skipped": skipped}

    @staticmethod
    def import_products(file_path):

        workbook = load_workbook(ImportService.resolve_file_path(file_path))

        print("\n========== PRODUCT IMPORT ==========")
        print("Sheets Found:\n")

        for sheet in workbook.sheetnames:
            print(f" - {sheet}")

        created = 0
        skipped = 0

        for sheet in workbook.sheetnames:

            supplier_name = sheet.strip()
            print(f"\nLooking for supplier: {supplier_name}")

            supplier = Supplier.query.filter_by(name=supplier_name).first()

            if supplier:
                print("✓ Supplier Found")
            else:
                supplier = Supplier(name=supplier_name)
                db.session.add(supplier)
                db.session.flush()
                print("✓ Supplier Created")

            worksheet = workbook[sheet]

            for row in worksheet.iter_rows(min_row=2, values_only=True):

                if not row:
                    continue

                brand = "" if row[0] is None else str(row[0]).strip()
                variant = "" if row[1] is None else str(row[1]).strip()

                vat_value = ""
                if len(row) > 2 and row[2] is not None:
                    vat_value = str(row[2]).strip().upper()

                selling_price = 0
                if len(row) > 3 and row[3] is not None:
                    selling_price = float(row[3])

                if not brand or not variant:
                    continue

                display_name = f"{brand} {variant}".strip()

                product = Product.query.filter_by(
                    display_name=display_name
                ).first()

                vat_enabled = vat_value in ["YES", "Y", "TRUE", "1"]

                vat_rate = 16 if vat_enabled else 0

                if product:

                    product.supplier_id = supplier.id
                    product.brand = brand
                    product.variant = variant
                    product.display_name = display_name
                    product.selling_price = selling_price
                    product.vat_enabled = vat_enabled
                    product.vat_rate = vat_rate

                    skipped += 1
                    continue

                new_product = Product(
                    supplier_id=supplier.id,
                    brand=brand,
                    variant=variant,
                    display_name=display_name,
                    selling_price=selling_price,
                    vat_enabled=vat_enabled,
                    vat_rate=vat_rate,
                )

                db.session.add(new_product)

                created += 1

        db.session.commit()

        print("\n========== IMPORT COMPLETE ==========")

        return {
            "created": created,
            "updated": skipped,
        }